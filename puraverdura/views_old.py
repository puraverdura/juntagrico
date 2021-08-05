# import vobject

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User
from django.utils.safestring import mark_safe
from juntagrico.dao.extrasubscriptioncategorydao import ExtraSubscriptionCategoryDao
from juntagrico.dao.memberdao import MemberDao
from juntagrico.dao.subscriptiondao import SubscriptionDao
from juntagrico.dao.subscriptionproductdao import SubscriptionProductDao
from juntagrico.dao.subscriptiontypedao import SubscriptionTypeDao
from juntagrico.entity.jobs import ActivityArea
from openpyxl import Workbook

import base64
import hmac
import hashlib
from urllib import parse
from django.conf import settings


from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render

from juntagrico.models import Member, Subscription

from juntagrico.dao.depotdao import DepotDao
from juntagrico.dao.listmessagedao import ListMessageDao
from juntagrico.dao.subscriptionsizedao import SubscriptionSizeDao
from juntagrico.util.pdf import render_to_pdf_http
from juntagrico.util.temporal import weekdays, start_of_business_year, end_of_business_year
from juntagrico.views import get_menu_dict
from juntagrico.config import Config
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl.utils import get_column_letter

from puraverdura.utils.stats import assignments_by_subscription, assignments_by_day, slots_by_day, \
    members_with_assignments, members_with_assignments_no_filter
from puraverdura.utils.utils import date_from_get, get_delivery_dates_of_month

from puraverdura.forms import MemberProfileForm

# API

# @staff_member_required
# def api_emaillist(request):
#     """prints comma separated list of member emails"""
#     sep = request.GET.get('sep', ', ')
#     format = request.GET.get('format', 'plain')
#     emails = sep.join(Member.objects.filter(inactive=False).values_list('email', flat=True))
#     if format == 'plain':
#         # just display for copy
#         return HttpResponse(emails)
#     else:
#         # create file for download
#         if format.find(';') == -1:
#             content_type = f'text/{format}'
#             file_type = format
#         else:
#             content_type, file_type = format.split(';')
#         response = HttpResponse(content_type=content_type)
#         response['Content-Disposition'] = f'attachment; filename="emails.{file_type}"'
#         response.write(emails)
#         return response


# @staff_member_required
# def api_vcf_contacts(request):
#     members = Member.objects.filter(inactive=False)
#     cards = []
#     for member in members:
#         card = vobject.vCard()

#         attr = card.add('n')
#         attr.value = vobject.vcard.Name(family=member.last_name, given=member.first_name)

#         attr = card.add('fn')
#         attr.value = member.get_name()

#         attr = card.add('email')
#         attr.value = member.email

#         if member.mobile_phone:
#             attr = card.add('tel')
#             attr.value = member.mobile_phone
#             attr.type_param = 'cell'

#         if member.phone:
#             attr = card.add('tel')
#             attr.value = member.phone
#             attr.type_param = 'home'

#         attr = card.add('org')
#         attr.value = ["meh als gmües"]
#         cards.append(card)

#     response = HttpResponse(content_type='text/x-vcard')
#     response['Content-Disposition'] = 'attachment; filename="members.vcf"'
#     response.write("".join([card.serialize() for card in cards]))
#     return response


@login_required
def nextcloud_profile(request):
    member = request.user.member
    membergroups = request.user.groups.values_list('name', flat=True)
    grouplist = list(membergroups)
    response = JsonResponse({'id': member.id,
                             'email': member.email,
                             'firstName': member.first_name,
                             'lastName': member.last_name,
                             'displayName': member.first_name + " " + member.last_name,
                             'roles': grouplist})
    return response


# pdf
def generate_pdf_dict():
    return {
        'subscriptions': SubscriptionDao.all_active_subscritions(),
        'products': SubscriptionProductDao.get_all(),
        'extra_sub_categories': ExtraSubscriptionCategoryDao.categories_for_depot_list_ordered(),
        'depots': DepotDao.all_depots_order_by_code(),
        'weekdays': {weekdays[weekday['weekday']]: weekday['weekday'] for weekday in
                     DepotDao.distinct_weekdays()},
        'messages': ListMessageDao.all_active()
    }


def other_recipients_names_w_linebreaks(self):
    members = self.recipients.exclude(email=self.primary_member.email)
    members = [str(member) for member in members]
    return mark_safe('<br>'.join([', '.join(members[x:x+6]) for x in range(0, len(members), 6)]))


@staff_member_required
def depot_list(request):
    depot_dict = generate_pdf_dict()
    for depot in depot_dict['depots']:
        depot.delivery_dates = list(get_delivery_dates_of_month(depot.weekday, int(request.GET.get('month', 0))))
    Subscription.other_recipients_names_w_linebreaks = other_recipients_names_w_linebreaks
    return render_to_pdf_http('exports/depotlist.html', depot_dict, 'depotlist.pdf')


@staff_member_required
def depot_overview(request):
    return render_to_pdf_http('exports/depot_overview.html', generate_pdf_dict(), 'depot_overview.pdf')


@staff_member_required
def amount_overview(request):
    return render_to_pdf_http('exports/amount_overview.html', generate_pdf_dict(), 'amount_overview.pdf')


#@permission_required('juntagrico.can_filter_members')
def filters_emails(request):
    members = MemberDao.active_members()

    renderdict = get_menu_dict(request)
    renderdict.update({
        'members': members
    })
    
    return render(request, 'members_only_emails.html', renderdict)

# tutorial webpage
@login_required
def tutorials(request):
    renderdict = get_menu_dict(request)
    return render(request, 'tutorials.html', renderdict)


@staff_member_required
def stats(request):
    activity_area = ActivityArea.objects.filter(pk=request.GET.get('activity_area', None)).first()

    start_date = date_from_get(request, 'start_date', start_of_business_year())
    end_date = date_from_get(request, 'end_date', end_of_business_year())

    filename = '{}_{}_{}stats.xlsx'.format(
        start_date.strftime('%Y-%m-%d'),
        end_date.strftime('%Y-%m-%d'),
        str(activity_area.pk) + '_' if activity_area else ''
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()


    # Sheet 5: master sheet
    ws5 = wb.active
    ws5.title = "master"

    # header
    ws5.cell(1, 1, u"{}".format(_('Mitgliedernummer')))
    ws5.column_dimensions['B'].width = 17

    ws5.cell(1, 2, u"{}".format('Vorname'))
    ws5.column_dimensions['C'].width = 17

    ws5.cell(1, 3, u"{}".format(_('Name')))
    ws5.column_dimensions['D'].width = 17

    ws5.cell(1, 4, u"{}".format(_('Strasse')))
    ws5.column_dimensions['E'].width = 17

    ws5.cell(1, 5, u"{}".format('PLZ'))
    ws5.column_dimensions['F'].width = 17

    ws5.cell(1, 6, u"{}".format(_('Mail')))
    ws5.column_dimensions['G'].width = 40

    ws5.cell(1, 7, u"{}".format('Telefon'))
    ws5.column_dimensions['H'].width = 17

    ws5.cell(1, 8, u"{}".format(_('Anmeldedatum')))
    ws5.column_dimensions['M'].width = 40

    ws5.cell(1, 9, u"{}".format(_('Depot')))
    ws5.column_dimensions['I'].width = 17

    ws5.cell(1, 10, u"{}".format(_('Arbeitseinsätze')))
    ws5.column_dimensions['J'].width = 17

    ws5.cell(1, 11, u"{}".format(_('Arbeitseinsätze Kernbereich')))
    ws5.column_dimensions['K'].width = 40

    ws5.cell(1, 12, u"{}".format('Mit-Abo-Mails'))
    ws5.column_dimensions['L'].width = 80

    ws5.cell(1, 13, u"{}".format(_('Anzahl Anteilscheine')))
    ws5.column_dimensions['N'].width = 40

    ws5.cell(1, 14, u"{}".format('Nr.Anteilschein(e)'))
    ws5.column_dimensions['A'].width = 17

    ws5.cell(1, 15, u"{}".format('Datum Einzahlung Anteilschein(e)'))
    ws5.column_dimensions['P'].width = 40

    ws5.cell(1, 16, u"{}".format(('Anteilschein(e) ausgestellt')))
    ws5.column_dimensions['Q'].width = 40

    ws5.cell(1, 17, u"{}".format(('Anzahl Ernteanteile')))
    ws5.column_dimensions['O'].width = 40

    #ws5.cell(1, 18, u"{}".format(('Datum Einzahlung Ernteanteil(e)')))
    ws5.cell(1, 18, u"{}".format(('Grösse Ernteanteil(e)')))
    ws5.column_dimensions['R'].width = 40

    # data
    members = members_with_assignments_no_filter(start_date, end_date)
    #members = MemberDao.all_members().filter(inactive=False)

    #users = User.objects.all()
    #print(users)

    for curr_row, member in enumerate(members, 2):
        curr_user = member.user
        
        ws5.cell(curr_row, 1, member.id)
        ws5.cell(curr_row, 2, member.first_name)
        ws5.cell(curr_row, 3, member.last_name)
        ws5.cell(curr_row, 4, member.addr_street)
        ws5.cell(curr_row, 5, member.addr_zipcode)
        ws5.cell(curr_row, 6, member.email)
        ws5.cell(curr_row, 7, member.phone)
        ws5.cell(curr_row, 8, curr_user.date_joined.date())

        sub = member.subscription
        depot = sub.depot.name if sub is not None else '-'
        ws5.cell(curr_row, 9, depot)

        assignments = member.assignments if member.assignments is not None else '-'
        ws5.cell(curr_row, 10, assignments)
        ws5.cell(curr_row, 11, 'N.A.')

        email_string = '-'
        if sub is not None:
            co_members = sub.other_recipients()
            email_list = [m.email for m in co_members]
            if email_list: 
                email_string = ', '.join(email_list)
        ws5.cell(curr_row, 12, email_string)

        curr_shares = member.active_shares
        num_shares = len(curr_shares)
        ws5.cell(curr_row, 13, num_shares)

        share_id_string = '-'
        share_ids = [str(s.id) for s in curr_shares]
        if share_ids:
            share_id_string = ', '.join(share_ids)
        ws5.cell(curr_row, 14, share_id_string)

        share_paid_date_string = '-'
        share_paid_dates = [str(s.paid_date) for s in curr_shares]
        if share_paid_dates:
            share_paid_date_string = ', '.join(share_paid_dates)
        ws5.cell(curr_row, 15, share_paid_date_string)

        share_issue_date_string = '-'
        share_issue_dates = [str(s.issue_date) for s in curr_shares]
        if share_issue_dates:
            share_issue_date_string = ', '.join(share_issue_dates)
        ws5.cell(curr_row, 16, share_issue_date_string)
        
        amount_subs_string = '-'
        size_subs_string = '-'
        if sub is not None:
            amount_subs_string = str(sub.subscription_amount(0))    
            size_subs_string = str(sub.size_name)    
        ws5.cell(curr_row, 17, amount_subs_string)
        ws5.cell(curr_row, 18, size_subs_string)

       


    # Sheet 1: assignments by subscription
    ws1 = wb.create_sheet(title="assignments by subscription")

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws1.column_dimensions['B'].width = 17
    ws1.cell(1, 3, u"{}".format(_('{}-Grösse').format(Config.vocabulary('subscription'))))
    ws1.column_dimensions['C'].width = 17

    # data
    for row, subscription in enumerate(assignments_by_subscription(start_date, end_date, activity_area), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription['subscription'].members.all()]))
        ws1.cell(row, 2, subscription['assignments'])
        ws1.cell(row, 3, subscription['subscription'].totalsize)

    # Sheet 2: assignments per day
    ws2 = wb.create_sheet(title="assignments per day")

    # header
    ws2.cell(1, 1, u"{}".format(_('Datum')))
    ws2.column_dimensions['A'].width = 20
    ws2.cell(1, 2, u"{}".format(_('Arbeitseinsätze geleistet')))
    ws2.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(assignments_by_day(start_date, end_date, activity_area), 2):
        ws2.cell(row, 1, assignment['day'])
        ws2.cell(row, 2, assignment['count'])

    # Sheet 3: slots by day
    ws3 = wb.create_sheet(title="slots per day")

    # header
    ws3.cell(1, 1, u"{}".format(_('Datum')))
    ws3.column_dimensions['A'].width = 20
    ws3.cell(1, 2, u"{}".format(_('Arbeitseinsätze ausgeschrieben')))
    ws3.column_dimensions['B'].width = 17

    # data
    for row, assignment in enumerate(slots_by_day(start_date, end_date, activity_area), 2):
        ws3.cell(row, 1, assignment['day'])
        ws3.cell(row, 2, assignment['available'])

    # Sheet 4: assignments per member
    ws4 = wb.create_sheet(title="assignments per member")

    # header
    ws4.cell(1, 1, u"{}".format(Config.vocabulary('member')))
    ws4.column_dimensions['A'].width = 40
    ws4.cell(1, 2, u"{}".format(_('Arbeitseinsätze')))
    ws4.column_dimensions['B'].width = 17

    # data
    members = members_with_assignments(start_date, end_date, activity_area)
    for row, member in enumerate(members, 2):
        ws4.cell(row, 1, u"{}".format(member))
        ws4.cell(row, 2, member.assignments)



    wb.save(response)
    return response



@login_required
def profile(request):
    success = False
    member = request.user.member
    if request.method == 'POST':
        memberform = MemberProfileForm(request.POST, instance=member)
        if memberform.is_valid():
            # set all fields of user
            member.first_name = memberform.cleaned_data['first_name']
            member.last_name = memberform.cleaned_data['last_name']
            member.email = memberform.cleaned_data['email']
            member.addr_street = memberform.cleaned_data['addr_street']
            member.addr_zipcode = memberform.cleaned_data['addr_zipcode']
            member.addr_location = memberform.cleaned_data['addr_location']
            member.phone = memberform.cleaned_data['phone']
            member.mobile_phone = memberform.cleaned_data['mobile_phone']
            member.iban = memberform.cleaned_data['iban']
            member.reachable_by_email = memberform.cleaned_data['reachable_by_email']
            member.save()
            success = True
    else:
        memberform = MemberProfileForm(instance=member)
    renderdict = get_menu_dict(request)
    renderdict.update({
        'memberform': memberform,
        'success': success,
        'member': member,
        'menu': {'personalInfo': 'active'},
    })
    return render(request, 'profile.html', renderdict)